# -*- coding: utf-8 -*-
"""
genera_dati_3lpp.py - Sakarya 3LPP Inspection - Weekly Summary

Legge il file "3LPP Sakarya Inspection Overall Status as of *.xlsm" piu'
recente presente in questa cartella e genera 3lpp.html: una pagina statica
autosufficiente (numeri gia' calcolati dentro l'HTML, nessun fetch/JS
esterno) con i KPI salienti, pronta per essere pubblicata insieme al resto
del sito (index.html ha un link "3LPP Inspection ->" in testata). La pagina
pubblicata e' in INGLESE (pubblico Sumitomo/Saipem); questo script e i log
restano in italiano per Rino.

Fogli letti (nomi esatti nel workbook):
  Production, Lab Tests PPT, Lab Tests Prod, On hold,
  Stripping- heating & Steel Dam.

NOTE METODOLOGICHE:
- Nei fogli "Lab Tests PPT/Prod" le celle di misura (col. H in poi)
  contengono per lo piu' il testo "X" (placeholder): l'esito Pass/Fail e'
  segnalato dal COLORE della cella (retino rosso = FAILED, giallo pieno =
  IN PROGRESS). Le righe con Pipe N° vuoto o "-" sono righe-modello/
  separatori e vengono escluse dal conteggio.
- "Non Approvate/Quarantena" viene dal log grezzo del foglio "On hold"
  (NON dalla tabella riepilogativa manuale a fianco, che puo' restare
  disallineata).

Avvio: doppio clic su AVVIA_GUI_3LPP.bat (GUI) oppure
       python genera_dati_3lpp.py  (da riga di comando)
"""
import glob
import math
import os
import re
import sys
from datetime import datetime

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
PATTERN = "3LPP Sakarya Inspection Overall Status as of *.xlsm"

# Logo IQS (stesso identico a quello nel piè di pagina di index.html)
IQS_LOGO_B64 = "data:image/jpeg;base64,/9j/4AAQSkZJRgABAQEAYABgAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoMDAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsNFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wAARCABXAMkDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD9U6KKKACikzTMgUAOwG96T73Wvn3xB+114X0348aN8NbJDqF5PNJa32oJJths7jb+7g/23d8Jx9zd3+fZ6/daxdxrI8SbM/d8ymc9LEUq3PyS+E6aiuFk8eSacESdY5pP7ufKLfnV/wAM/EHRfFF5NZW1x5OoQ/fs5vkf/wCypHQdXRRRQAUUUUAFFFFABRSUUALRRRQAUUUUAFFFFABRSUtABRRRQAUhpaSgCrIyW8byu6qiLlmY8Cvi79qj9qK91P4VXF98KdbjbT7fV30rWdWtPluYTs3wmH/pk/zp53+x8ma+05Yo5oXjdVeNxtZXHBr83PFfwhPwt+OfjH4Zwrs8K+OrB002OQ4jSb/XWf8A3xcJ5P8AuOKo+azrEVcPSiofDL3f8jzD496tJf8AijwR8WNHjW0ufEllDqUkiYSOLVLZ/JufydEb/gVfpm/jjSv+EHtPFV1fW+maLcWMN+91eypHDFC6I/zu/wDv1+W0DDXP2c54GXfceHfEaPEz/wAEN3bPvH/fdotJqXi3Vfif4Qsr/wCIt7cQ/CzwLFbaZp2hadL5L6zf+T8kO/8Av7Pnd/4E+58710YWh7aZ4GRYr/aJr+aPN/8AJH1H4u/bW8H65rE9p4M0rxJ4ye0/4+L3QtO/0aH/AG3eZ0/+Iq14L+M2j+OLz7Ra6nYpqlv/AKmH7XC+oJ/wCF3r491+31PVIdEtPFGmPdT32ybQPhdoW+2s7dH+5Nconz/P9/8A57P993SvZfCP7JOt/Y4LvxbraeF/400LwdDDZ+T/ALDzJ9//AMf/AN+vTxGFoQh8R9vCrVmffXwp+JH/AAlcL6ffsqatbpu3f890/v16PXxX4Hjl8A6pZT2V7fXz2779+oXDzO//AAN6+y7O6ivrOG5gO6GZFdP9014R2n55/tgft5fEv4FfHbWvCHhu38Py6RaW9tJE2o2Uk0uXhR3+ZZk/v1k/stf8FBvih8ZPj54R8G+ILbw9Fo+qyzJcNY2MyS4S2mf5G85x99Erwr/gpVx+1x4n/wCvSw/9Jkrn/wBgH/k8L4bf9fN5/wCkVzX2EMJh/qXtOTXlPG9tP23Ift92ri/il8VvDHwd8H3fiTxbqkOk6Zb/AC+Yx3PI5+6iL1dz/dFdLqurWuhaZd6hfzJbWdpE09xM/CRoilmc+wFfhx+1b+0jq/7SnxOutYnlmg8N2TvDo2mP9y3h/vun99/vv/3x/BXg4HBPGTt9k78RW9jG59E/F/8A4KueLdau57T4c6HaeHdLzsTUNXT7ReN/t7P9Sn+589eB6j+3T8d9Um824+JWpo//AE6w20Kf98IiV5L4N8E6/wDEbxJZ+HvDOlz6zrN422Ozt1+dv/iE/wBt6+z/AAb/AMElfG+raalx4i8Z6T4eunXeLW1tXvNv+w770Qf8B3V9PKngMHpM8znxFbY8g8L/APBQr48eGp0L+Nf7WgT79rqtlbyo3/Awm/8A8fr7H/Z3/wCCoHh7x9qFpoPxF0+LwdqlwdkWrwyF9NlbP8e/54f/AB9PV1r5V+N//BOf4m/B3RrjXLJrTxpo1su+dtJ3/aok/vvC/wDD/ub6+VaPqeDxsP3Qe2rUZe+f0fRyLIu5W3rUnFfnP/wTG/akvNcYfCHxRdPcz2lu83h66mfD+Sn37M/7ifOn+wj9kWvt741eNJvhx8I/Gvie22tc6Po13fQI44aWOF3QfiwWvj6+HnRq+xZ68K0Zw50fMn7W/wDwUK0z4G6zdeEfBtnb+JfF8Hy3U1zKfsWnP/ck2fO7/wCx8m3P3/4K+HNX/wCChXx81a7kn/4T19Pj3/Ja2emWion/AJB3/wDfdfPl9fXGq311e3dw91dXEzzTTTPveZ3+d3evrL9iT9iLTf2mtA1rxL4h8Q3mmaNp96dNSz0jZ9plmCI7vvdHRE+dP4H3/P8A3K+t+p4TB0eepqeR7arWn7hm+Ef+Cl/x08N3CtqGt6b4nhIx5OraZCn/AI9b+S1foT+x/wDtXT/tSeG9WvZ/CV14ek0qZIJrqO6Saznd0ztRzsff3KFPl3J8xzXzV8Wv+CSsNrpM118NfFt3dXsS7l0zxGU/en/YmhRNn/fH/A6+1v2dfg3Y/Af4R6H4OstpltIt17crkfaLl+ZZOnTd930UKK8jG1cFOl+4h7x24eFWMvfPUaWiivAPQCiikoAP4a+bf2xPCr3Fj4S8WWqY1DRNQ/dsn+186f8Aj8SD/gVfSNcF8atLTWvAN9bsm/54mX67xTOHG0vbUJQPzh8TeHP7Lh+OOmW6bLJNRhe3j/2E1B0T/wAcesrQ9Mtdb1zwXpurWaz+G/Beh3PiC+spPuXcz77n5/8Af/0OH/cSvbPHnhV5E+Izov7y71SG2/8AH5n/APaNc7pXhz7UmttFF5E+p+GEstn9x7byd/8A45bb/wDgdetgZfEfI5TgfY15z/r4jr/2UPhldajYav8AFjxA6aj4r167m+z3V0n+ph37HdP7m996f7iV674j3733pUngfZ4f+C3hqK0ldHSxRNn+3/HXlfirXNQtXfyr2ZP+B1xYiXPVkfcxjyRNiT/j5r6p+Gt0ZvA+js3XyvL/AO+GKf0r4+8P+IH1z/Wp5c6/f/26+yfANidP8G6PGxy/2ff/AN9/P/WuUs/IL/gpX/ydx4n/AOvSw/8ASZK579gH/k8L4bf9fN5/6RXNdD/wUr/5O48T/wDXpYf+kyVz37AP/J4Xw2/6+bz/ANIrmvvof8i7/t0+e/5iP+3j9Kv+ChPi648J/sneMpLSVoJ9Q+z6YZIz/BNOiS/mm9fxr8U6/an/AIKG+FrnxR+yb4y+zKZJ9NNvqYT/AGIZkL/+Ob2/CvxWrhyb+A7HRjfiP1p/4Jg/BnTfB/wTTx5JAsniHxRLMPtDqN0VpFO6JEvszo7/AIp/cFfa618W/wDBMP4yab4v+BqeCWmWPXfC80yvbs/zzW80zypMo9N7un/AF/vV6x+1F+0pqn7N/h618QReBLzxZorN5dzfWt6sKWb4+Tzso21GyBv+v+xu8DFQq1cVOLO+jKMKSPecA1+MH/BRL4L6d8H/ANoSZ9Dt1stJ8QWiasltCoRIJnd0mVPq6b/+B17jD/wV41GbXbPzvhzb2mjeav2rZqbTXIh/j2fIib6+fv29Pj/ov7QXxntdS8MTPd+H9N0yGzt7l0dPNf55nfY4+T/XbP8AgFepl+FxGGxC50cuIrUpw908u/Z/8WXHgX44eBNct5njey1mzd/L/jh85EdP+Bo7pX7l/GTwTJ8SfhP4v8KwsiS61pN1YRvJ91JJInVG/Bitfhp8A/Cc/jn43eBNDgR5Hvdcs0fy/wCBPOR3f/gCI71/QD2ozmXLWhJbhgfhkfzma5oeoeGdavdH1W0m0/VLGZ7W4tZk2PE6ffR67D4S/Hbxx8DdWnvvBXiG40Z7jZ9oh+Sa2n92if5GNfsr8cP2P/hh8fZvtnibQTDrO3Yms6ZMbe5A/wBr+B/+Bo9fG/xQ/wCCSWq2Mc114A8Zx6ntO+PTddh8p/8Av8nyO/8AwBK7aOZ4etD2ddGM8LVhrAi+GP8AwVs16xmgt/H/AIOtNTtc7H1Dw+7xTL/t+S+9H/77Sv0A+D/xo8I/HLwmniHwfq8Wp2W7ZKn3ZbeT+5KnVH9q/BDxX4U1fwH4l1Pw9r1jJpus2EvlXVvLyytX0j/wTa+ImoeDf2oNF0e3uG/szxNFNYXsH8D7IXmhf/fR0x/20escdllFUfb0C6OInz8kz9nKKKK+TPXCiiigBq9axvFFqt9pv2d/9W7fP/u1sr1qvfQfaoNn96gmWx8/Q+AP7c1TTre4i/4+759QuP8Ac/yj/wDfdbXxC+EEV1q767o6ol0775rX+B3/AI3r1Ox0dLW5kutuXZdif7lNvqulOUPgOelR5DwODTZdN0F9NeKZPs7u6I/8CV5D4x++9fS/jy6/cvbxRefdP/45XmOn/Bu/8XagVki/d/x/N8i/770SlznUcj8AvBdx4w8WJF5X/Euh+e7f/Y/uV9uKNi7a5vwL4F034e6KthpsWB9+abb88reprpqgD8X/APgpV/ydx4n/AOvSw/8ASZK5/wDYB/5PC+G3/Xzef+kVzX1l+11+wF8SPjx8dNY8ZeHtT8M2ukXdvbxqmqXdwk2UiRH+SOB/7n9+sj9l7/gnj8Tvgv8AHrwn4z1zU/C8+k6TLM9wmnX1y8zb7WaH5FeBP45P79fYQxmH+pez59eU8b2M/bc5+iWraRaa9pN7puoQpc2V7E9vcQv92VHUoyfiK/DP9qn9nLWP2a/iddaLdQzTeHrp3m0bU5PuXFt/c/30+4//AMQ6V+7vauI+LPwg8KfGnwjceHPF2lR6nps3zLuG2SB+zxPjKN714OBxrwc7/ZO/EUfbRPwT8C+Ptf8Ahp4os/EPhjVLjRtasm3xXNs//jj/AN9P9h6+ztD/AOCrPiK68OzaT4y+HOg+KkuIXguPIuXtobhM7X3xOkyvv9OnNO+MH/BKXxroF5PdfDvWrPxPpmd8On6hJ9kvU/2N/wDqX/3/AJP9yvn/AFL9ij45aTN5Vx8NdYd/+nVUuU/77R3r6eVTAYx882eXyYiieU+LdW0rXPE2oahomj/8I9pdxM72+l/aHufs6f3N7/fSsivovwl/wT6+O/i25Rf+EKbSIH+/datdwwon/AN+/wD8crR/aG/YD8d/AHwZpfiOW7g8V2jny9T/ALJhf/iXtnCYL/M6P/f2JhyPUV1rGYbn5Ocx9jP4+Q+gv+CW/wCzeIVl+L+trG/mrLY6JArlyuPlmuW/u9HRR6F/9ivvH4rfEnSPhH8PNc8X61J5en6TbtO6g4eZuiQp/tu5VB7tX5Q/sa+LPjt8EfFaXHh/4feLPEfhDUGT+0dK/sy4S2l/6bI+zYk3/of8dfp58XPhD4b/AGkvhiNA8U2mpWen3QivI445Tb3VpNsyh4yu9C5+V9657Gvkswj/ALXz1Ze6exh5fuvcPy/+GP8AwUg+KXgXx5rGsarJH4q0XVr57y40K8mdEt9/8FtMf9Ts+5s+dP8AY3/PX1En/BW74ff2N57eDvEw1fb/AMeuLfyN3/XXzN2332V4n8UP+CUfj/w/cS3HgjXNN8VacRlLW9b7Jef7nzfI/wDv70/3K8gm/wCCf3x/hufK/wCFd3En+3HqNns/9HV68qeW4j3rnFzYiB5n8bvitqHxy+KfiHxvqVrFZXWrSo/2WH7kSIiIif7fyInz179/wTH+Gt74y/aUstfWFzpfhe0mvLmb+DzpkeGFP9/53f8A4A9bnwv/AOCWPxO8T3UD+Lr7S/BmnbsyoJvtl5/wBU+T/wAfr9Jfgd8B/Cf7PvguLw34Tsngt93m3FzO2+4u5f78r9zU43HUY0fYUCsPh5ynzzPTKKSlr5M9cKKKKACiioJoWk6Sun+5QA+SIPVOfS1n/wCWtQXGjzzfd1O6j/3KoyeF7qT/AJjuo/8AfS0ATx+D9Mjl3tF5z/8ATRq2beGK1jEUKLGi/wAK1j2/h24h/wCYxfSf771pW9m8P3rqWT/fNAFuimU+gDgNQ8QeLYb67gt9GhmiR2hSTosrvnyX+99xPlWXj+P5PufNLZeKPEd9fC1h8NrGFTe811czRIr/ANw5h/VN61pzab4pkkdo9a0tFLfKr6S77f8AyZpf7P8AFn/Qf0n/AMFD/wDyTVAc7N408VQ3oiPh1ZI4rx1n2mbe1sqO+9Pk2b+EAXf8/wDs7/ktXHjLxLCsb/8ACJSxsJkS4VrnftR+d6bE+fYhXfj+MbE3/frY/s/xZ/0H9J/8FD//ACTR/Z/iz/oP6T/4KH/+SaAMHwv4u8QalrEdlqHh6SzieaUfbD5oTygHKP8AMg9EXD7G9quX3iTxJZwajOnh6O7W0Z/LRLhllnxI6LsTY/8AAEb/AIFWl/Z/iz/oP6T/AOCh/wD5Jo/s/wAWf9B/Sf8AwUP/APJNAHLw/ErxBNJdJH4LunmjWGSOFpHR5d7up/5Y7E27P43q1b+PPEl3bzFPCbpN92FZpJk87/yD8n/A9lb39n+LP+g/pP8A4KH/APkmj+z/ABZ/0H9J/wDBQ/8A8k0AYknjbX454kTwvJOjwb96GVMzbj+5+eHj/ffYnP3607bWPEFxo97cS6NHaagmnpc20C3Hmo0zK/7lvkT7jKv/AH3Vj+z/ABZ/0H9J/wDBQ/8A8k0f2f4s/wCg/pP/AIKH/wDkmgDDuPFuuQxyGLT5NVTyiySLYXFniTn938+9vn/vnYiY5erM3irxLb6S98fDEB22iXTQfbn83nO+L/U/fWtP+z/Fn/Qf0n/wUP8A/JNH9n+LP+g/pP8A4KH/APkmgDlpPiJr6zW/leDrq5SQA+bC8yIOV+XDwo38XXG3itNfFmsxahpdvd+Gwk98sZTybov5Tt94P8n8CK+5/XYv8da39n+LP+g/pP8A4KH/APkmj+z/ABZ/0H9J/wDBQ/8A8k0AdNRVe3RlhVZX8yTb877du6rFSAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAf/9k="

SH_PROD = "Production"
SH_LABPPT = "Lab Tests PPT"
SH_LABPROD = "Lab Tests Prod"
SH_ONHOLD = "On hold"
SH_STRIP = "Stripping- heating & Steel Dam."
SH_RAWMAT = "Raw Material"

FOGLI_RICHIESTI = [SH_PROD, SH_LABPPT, SH_LABPROD, SH_ONHOLD, SH_STRIP, SH_RAWMAT]


def log(msg):
    print(msg, flush=True)


def trova_file_excel(cartella=HERE):
    candidati = [f for f in glob.glob(os.path.join(cartella, PATTERN))
                 if not os.path.basename(f).startswith("~$")]
    if not candidati:
        return None

    def chiave_data(path):
        m = re.search(r"as of (\d{1,2})\.(\d{1,2})\.(\d{4})", os.path.basename(path))
        if m:
            mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return (yyyy, mm, dd)
        return (0, 0, 0)

    return max(candidati, key=lambda p: (chiave_data(p), os.path.getmtime(p)))


def data_da_nome_file(path):
    m = re.search(r"as of (\d{1,2})\.(\d{1,2})\.(\d{4})", os.path.basename(path))
    if m:
        mm, dd, yyyy = m.group(1), m.group(2), m.group(3)
        return "%02d/%02d/%s" % (int(dd), int(mm), yyyy)
    return datetime.now().strftime("%d/%m/%Y")


def num(v, default=0):
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    return default


def somma_colonna(ws, col, riga_ini, riga_fin):
    tot = 0
    for r in range(riga_ini, riga_fin + 1):
        tot += num(ws.cell(row=r, column=col).value)
    return tot


def conta_status(ws, col_status, riga_ini, riga_fin, valore=None):
    """Conta le righe con un valore in col_status; se valore=None conta
    tutte le righe non vuote (qualsiasi status)."""
    n = 0
    for r in range(riga_ini, riga_fin + 1):
        v = ws.cell(row=r, column=col_status).value
        if v is None or str(v).strip() == "":
            continue
        if valore is None or str(v).strip() == valore:
            n += 1
    return n


def ultima_riga(ws, col, riga_ini):
    last = riga_ini - 1
    for r in range(riga_ini, ws.max_row + 1):
        if ws.cell(row=r, column=col).value not in (None, ""):
            last = r
    return last


# ----------------------------------------------------------------------
#  Raw Material: 5 componenti (Manufacturer/Trade Name/Batch ripetuti in
#  colonne C:Q), una riga per data/turno. Si mostra lo stato ATTUALE
#  (ultima riga = ultima consegna/turno registrato) + un avviso se un
#  componente ha cambiato lotto durante il periodo (rilevante per
#  tracciabilita').
# ----------------------------------------------------------------------
def estrai_raw_material(ws):
    last = ultima_riga(ws, 1, 5)
    if last < 5:
        return [], None

    slots = [(3, 4, 5), (6, 7, 8), (9, 10, 11), (12, 13, 14), (15, 16, 17)]
    materials = []
    for mc, tc, bc in slots:
        manufacturer = str(ws.cell(row=last, column=mc).value or "").strip()
        trade_name = str(ws.cell(row=last, column=tc).value or "").strip()
        batch = str(ws.cell(row=last, column=bc).value or "").strip()
        distinct_batches = set()
        for r in range(5, last + 1):
            b = ws.cell(row=r, column=bc).value
            if b:
                distinct_batches.add(str(b).strip())
        materials.append(dict(
            manufacturer=manufacturer, trade_name=trade_name, batch=batch,
            batch_changed=len(distinct_batches) > 1,
        ))
    last_date = ws.cell(row=last, column=1).value
    return materials, last_date


# ----------------------------------------------------------------------
#  Lab Tests: scansione colore cella (Pass/Fail/In Progress)
# ----------------------------------------------------------------------
def scan_lab_tests(ws):
    last_row = ultima_riga(ws, 2, 12)   # col B = Pipe N°
    if last_row < 12:
        return 0, 0, 0, 0

    # ultima colonna non vuota nella riga di intestazione (riga 8)
    last_col = 7
    for c in range(ws.max_column, 7, -1):
        if ws.cell(row=8, column=c).value not in (None, ""):
            last_col = c
            break

    n_pass = n_fail = n_prog = n_tot = 0
    for r in range(12, last_row + 1):
        pipe_val = str(ws.cell(row=r, column=2).value or "").strip()
        if pipe_val == "" or pipe_val == "-":
            continue
        is_fail = False
        is_prog = False
        for c in range(8, last_col + 1):
            cell = ws.cell(row=r, column=c)
            fill = cell.fill
            if fill is None or fill.patternType is None:
                continue
            fg = fill.fgColor
            rgb = fg.rgb if fg is not None and fg.type == "rgb" else None
            if fill.patternType in ("lightUp", "solid") and rgb == "FFFF0000":
                is_fail = True
                break
            if fill.patternType == "solid" and rgb == "FFFFFF00":
                is_prog = True
        n_tot += 1
        if is_fail:
            n_fail += 1
        elif is_prog:
            n_prog += 1
        else:
            n_pass += 1
    return n_pass, n_fail, n_prog, n_tot


# ----------------------------------------------------------------------
#  Estrazione KPI completa
# ----------------------------------------------------------------------
def estrai_kpi(path):
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    mancanti = [s for s in FOGLI_RICHIESTI if s not in wb.sheetnames]
    if mancanti:
        raise ValueError("Fogli mancanti nel file: %s" % ", ".join(mancanti))

    wsProd = wb[SH_PROD]
    tot_pipes_to_finish = num(wsProd["E3"].value)
    tot_item_qty = num(wsProd["B4"].value)
    work_days_to_finish = num(wsProd["E4"].value)
    production_estimation = wsProd["I3"].value
    qcp_no = str(wsProd["D13"].value or "").strip()
    od_wt = str(wsProd["F13"].value or "").strip()
    # "Pipes to Finish" e "Total item Qty" sono due totali dichiarati fianco a
    # fianco nello stesso foglio (Production!E3/B4, stessa riga di intestazione
    # del vendor): completato = item qty totale - pipe ancora da finire.
    completed_qty = max(tot_item_qty - tot_pipes_to_finish, 0)
    pct_complete = (completed_qty / tot_item_qty) if tot_item_qty > 0 else 0

    r1_ini, r1_fin = 16, ultima_riga(wsProd, 2, 16)     # blocco VDI (col B = Date)
    r2_ini, r2_fin = 16, ultima_riga(wsProd, 14, 16)    # blocco Finale (col N = Date)

    coated_vdi = somma_colonna(wsProd, 4, r1_ini, r1_fin)     # D
    na_vdi = somma_colonna(wsProd, 5, r1_ini, r1_fin)         # E
    repair_vdi = somma_colonna(wsProd, 6, r1_ini, r1_fin)     # F
    app_vdi = somma_colonna(wsProd, 7, r1_ini, r1_fin)        # G
    coated_fin = somma_colonna(wsProd, 16, r2_ini, r2_fin)    # P
    na_fin = somma_colonna(wsProd, 17, r2_ini, r2_fin)        # Q
    repair_fin = somma_colonna(wsProd, 18, r2_ini, r2_fin)    # R
    app_fin = somma_colonna(wsProd, 19, r2_ini, r2_fin)       # S
    coated_totale = coated_vdi + coated_fin

    wsOH = wb[SH_ONHOLD]
    last_oh = ultima_riga(wsOH, 6, 5)
    on_hold_count = conta_status(wsOH, 6, 5, last_oh)
    # scomposizione per disposizione - somma sempre a on_hold_count (stesso
    # foglio/colonna, non fonti diverse): questa e' la ripartizione corretta
    # da mostrare, NON na_vdi/na_fin sotto (quelli vengono dal tally
    # giornaliero del foglio Production, un conteggio diverso che non deve
    # essere presentato come "componenti" di on_hold_count).
    oh_not_approved = conta_status(wsOH, 6, 5, last_oh, "Not Approved")
    oh_on_hold = conta_status(wsOH, 6, 5, last_oh, "On hold")
    oh_repair = conta_status(wsOH, 6, 5, last_oh, "Repair")

    # Riepilogo manuale del vendor, riga 2 del foglio On hold (A2:H2):
    # etichetta/valore alternati. Fonte DIVERSA dal log grezzo sopra (puo'
    # non coincidere - "Not Approved" qui e' spesso 0 mentre il log ne
    # conta di piu': e' un riepilogo a parte, non un totale alternativo).
    oh_summary_row2 = []
    for c in range(1, 9, 2):
        label = wsOH.cell(row=2, column=c).value
        value = wsOH.cell(row=2, column=c + 1).value
        if label:
            oh_summary_row2.append((str(label).strip(), num(value)))

    # After Action (On hold!H5:H..) - esito dopo l'azione presa sul difetto
    # (colonna E=Defect, F=Status, G=Work station, H=After Action, I=After
    # Action Date): Approved/Not Approved/ancora vuoto (azione non ancora
    # presa). E' un dato DIVERSO dalla colonna Status (F) sopra - F e' lo
    # stato corrente della non conformita', H e' l'esito dopo l'azione.
    oh_action_approved = 0
    oh_action_not_approved = 0
    oh_action_pending = 0
    for r in range(5, last_oh + 1):
        h = str(wsOH.cell(row=r, column=8).value or "").strip()
        e = wsOH.cell(row=r, column=5).value
        if not e:
            continue
        if h == "Approved":
            oh_action_approved += 1
        elif h == "Not Approved":
            oh_action_not_approved += 1
        else:
            oh_action_pending += 1

    wsST = wb[SH_STRIP]
    last_st = ultima_riga(wsST, 6, 5)
    strip_count = conta_status(wsST, 6, 5, last_st)

    raw_materials, raw_material_date = estrai_raw_material(wb[SH_RAWMAT])

    pass_ppt, fail_ppt, prog_ppt, tot_ppt = scan_lab_tests(wb[SH_LABPPT])
    pass_prod, fail_prod, prog_prod, tot_prod = scan_lab_tests(wb[SH_LABPROD])
    fail_lab_totale = fail_ppt + fail_prod
    tot_lab = tot_ppt + tot_prod

    pct_na = (on_hold_count / coated_totale) if coated_totale > 0 else 0
    pct_fail = (fail_lab_totale / tot_lab) if tot_lab > 0 else 0
    pct_app_fin = (app_fin / coated_fin) if coated_fin > 0 else 0

    return dict(
        tot_pipes_to_finish=tot_pipes_to_finish, tot_item_qty=tot_item_qty,
        work_days_to_finish=work_days_to_finish, production_estimation=production_estimation,
        qcp_no=qcp_no, od_wt=od_wt,
        completed_qty=completed_qty, pct_complete=pct_complete,
        coated_vdi=coated_vdi, coated_fin=coated_fin, coated_totale=coated_totale,
        na_vdi=na_vdi, na_fin=na_fin, repair_vdi=repair_vdi, repair_fin=repair_fin,
        app_vdi=app_vdi, app_fin=app_fin, pct_app_fin=pct_app_fin,
        on_hold_count=on_hold_count, pct_na=pct_na,
        oh_not_approved=oh_not_approved, oh_on_hold=oh_on_hold, oh_repair=oh_repair,
        oh_summary_row2=oh_summary_row2,
        oh_action_approved=oh_action_approved, oh_action_not_approved=oh_action_not_approved,
        oh_action_pending=oh_action_pending,
        raw_materials=raw_materials, raw_material_date=raw_material_date,
        strip_count=strip_count,
        pass_ppt=pass_ppt, fail_ppt=fail_ppt, prog_ppt=prog_ppt, tot_ppt=tot_ppt,
        pass_prod=pass_prod, fail_prod=fail_prod, prog_prod=prog_prod, tot_prod=tot_prod,
        fail_lab_totale=fail_lab_totale, tot_lab=tot_lab, pct_fail=pct_fail,
    )


# ----------------------------------------------------------------------
#  Rendering HTML (statico, numeri gia' incorporati - nessun JS/fetch)
#  Pagina pubblicata in INGLESE.
# ----------------------------------------------------------------------
def semaforo(v, s1, s2):
    if v < s1:
        return "#00e676"
    if v < s2:
        return "#ffca28"
    return "#ff4757"


def fmt_en(n):
    return "{:,.0f}".format(n)


def fmt_pct(v):
    return "%.1f%%" % (v * 100)


def fmt_date(v):
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    return str(v or "n/a")


def stacked_bar(segments):
    """segments: list of (label, value, color). Renders one thin stacked
    horizontal bar (2px gaps between segments, rounded ends) + a legend row
    with swatch, label, value and share - so identity is never color-alone."""
    total = sum(v for _, v, _ in segments) or 1
    bar_parts = []
    legend_parts = []
    n = len(segments)
    for i, (label, value, color) in enumerate(segments):
        pct = value / total * 100
        radius = ""
        if n == 1:
            radius = "border-radius:4px;"
        elif i == 0:
            radius = "border-radius:4px 0 0 4px;"
        elif i == n - 1:
            radius = "border-radius:0 4px 4px 0;"
        bar_parts.append(
            '<div class="seg" style="width:%.3f%%;background:%s;%s" '
            'title="%s: %s (%.1f%%)"></div>' % (pct, color, radius, label, fmt_en(value), pct))
        legend_parts.append(
            '<div class="legend-item"><span class="dot" style="background:%s;"></span>'
            '%s <b>%s</b><span class="lg-pct">%.1f%%</span></div>' % (color, label, fmt_en(value), pct))
    bar_html = '<div class="stackbar">%s</div>' % "".join(bar_parts)
    legend_html = '<div class="legend">%s</div>' % "".join(legend_parts)
    return bar_html + legend_html


def genera_html(kpi, report_date):
    col_na = semaforo(kpi["pct_na"], 0.03, 0.07)
    col_fail = semaforo(kpi["pct_fail"], 0.02, 0.05)

    ring_r = 38
    circumference = 2 * math.pi * ring_r
    ring_offset = circumference * (1 - kpi["pct_complete"])

    bullets = []
    bullets.append("Production: %s pipes coated (VDI+Final) out of %s total order item qty; %s pipes still to finish (%s work days)." % (
        fmt_en(kpi["coated_totale"]), fmt_en(kpi["tot_item_qty"]), fmt_en(kpi["tot_pipes_to_finish"]), fmt_en(kpi["work_days_to_finish"])))
    bullets.append("Repair: %s pipes reworked at VDI, %s at Final Inspection (%s combined)." % (
        fmt_en(kpi["repair_vdi"]), fmt_en(kpi["repair_fin"]), fmt_en(kpi["repair_vdi"] + kpi["repair_fin"])))
    bullets.append("Anti Corrosive Plant: %s pipes not approved / quarantined (%s of total coated)." % (
        fmt_en(kpi["on_hold_count"]), fmt_pct(kpi["pct_na"])))
    bullets.append("Laboratory: %s tests performed (PPT %s / Prod %s), %s failed (%s), %s in progress." % (
        fmt_en(kpi["tot_lab"]), fmt_en(kpi["tot_ppt"]), fmt_en(kpi["tot_prod"]),
        fmt_en(kpi["fail_lab_totale"]), fmt_pct(kpi["pct_fail"]), fmt_en(kpi["prog_ppt"] + kpi["prog_prod"])))
    bullets.append("Stripping Control: %s non-conformances recorded." % fmt_en(kpi["strip_count"]))
    if kpi["pct_na"] > 0.07 or kpi["pct_fail"] > 0.05:
        bullets.append("WARNING: scrap/quarantine above reference threshold - investigate root causes.")

    bullets_html = "\n".join('<li>%s</li>' % b for b in bullets)

    # Ruolo di ciascuno dei 5 componenti (colonne C:Q di Raw Material),
    # confermato da Rino - il foglio stesso non li etichetta.
    RM_ROLES = ["Epoxy Powder", "Adhesive", "Powder Adhesive", "PP Top Coat", "Rough Coat"]
    rm_rows = []
    for i, m in enumerate(kpi["raw_materials"]):
        role = RM_ROLES[i] if i < len(RM_ROLES) else "Material %d" % (i + 1)
        badge = '<span class="rm-badge">BATCH CHANGED THIS PERIOD</span>' if m["batch_changed"] else ""
        rm_rows.append(
            '<tr><td>%s</td><td>%s</td><td>%s</td><td>%s%s</td></tr>' % (
                role, m["manufacturer"] or "n/a", m["trade_name"] or "n/a", m["batch"] or "n/a", badge))
    rm_table_rows = "\n".join(rm_rows)
    rm_date = fmt_date(kpi["raw_material_date"])

    chart_action = stacked_bar([
        ("Approved", kpi["oh_action_approved"], "#00e676"),
        ("Not Approved", kpi["oh_action_not_approved"], "#ff4757"),
        ("Pending Action", kpi["oh_action_pending"], "#ffca28"),
    ])

    # ---- Charts (thin stacked bars, verified compositions only) ----
    # Colori come var(--blue/--teal/--purple), non esadecimale fisso: cosi'
    # la palette accento scelta nei Tweaks del dashboard principale (ocean/
    # ember/signal, ereditata anche qui) ricolora anche questi grafici. I
    # colori di stato (Pass/Fail/In Progress sotto) restano fissi: sono
    # semantici, non devono cambiare con l'accento.
    chart_completion = stacked_bar([
        ("Completed", kpi["completed_qty"], "var(--teal)"),
        ("Remaining", kpi["tot_pipes_to_finish"], "var(--blue)"),
    ])
    chart_coated = stacked_bar([
        ("VDI Station", kpi["coated_vdi"], "var(--blue)"),
        ("Final Station", kpi["coated_fin"], "var(--teal)"),
    ])
    chart_quarantine = stacked_bar([
        ("Not Approved", kpi["oh_not_approved"], "var(--blue)"),
        ("On Hold", kpi["oh_on_hold"], "var(--teal)"),
        ("Repair", kpi["oh_repair"], "var(--purple)"),
    ])
    # Riepilogo manuale del vendor (On hold!A2:H2) - card a parte, valori
    # cosi' come compilati nel foglio, NON ricalcolati (possono differire
    # dal conteggio del log grezzo sopra, vedi nota in fondo alla pagina).
    oh_summary_html = "".join(
        '<div class="stat-item"><div class="stat-label">%s</div><div class="stat-val">%s</div></div>' % (
            label, fmt_en(value))
        for label, value in kpi["oh_summary_row2"]
    )

    # Disposizione per stazione (Approved/Repair/Not Approved), dal foglio
    # Production stesso - colori di STATO (fissi, non legati all'accento),
    # diverso concetto da "Not Approved/Quarantine by Disposition" sopra
    # (quello viene dal log del foglio On hold, un'altra fonte).
    station_rows = []
    for name, coated, app, rep, na in (
            ("VDI Station", kpi["coated_vdi"], kpi["app_vdi"], kpi["repair_vdi"], kpi["na_vdi"]),
            ("Final Station", kpi["coated_fin"], kpi["app_fin"], kpi["repair_fin"], kpi["na_fin"])):
        segs = [("Approved", app, "#00e676"), ("Repair", rep, "#ffca28"), ("Not Approved", na, "#ff4757")]
        total = max(coated, 1)
        bar = "".join(
            '<div class="seg" style="width:%.3f%%;background:%s;" title="%s: %s"></div>' % (
                v / total * 100, color, label, fmt_en(v))
            for label, v, color in segs if v > 0
        )
        station_rows.append(
            '<div class="lab-row"><div class="lab-row-lbl">%s <span class="lab-row-n">(%s coated)</span></div>'
            '<div class="stackbar">%s</div></div>' % (name, fmt_en(coated), bar))
    chart_station_rows = "\n".join(station_rows)
    chart_station_legend = '<div class="legend">' + "".join(
        '<div class="legend-item"><span class="dot" style="background:%s;"></span>%s</div>' % (c, l)
        for l, c in (("Approved", "#00e676"), ("Repair", "#ffca28"), ("Not Approved", "#ff4757"))
    ) + "</div>"

    lab_rows = []
    for name, p, f, prog, tot in (
            ("PPT", kpi["pass_ppt"], kpi["fail_ppt"], kpi["prog_ppt"], kpi["tot_ppt"]),
            ("Prod", kpi["pass_prod"], kpi["fail_prod"], kpi["prog_prod"], kpi["tot_prod"])):
        segs = [("Pass", p, "#00e676"), ("Fail", f, "#ff4757"), ("In Progress", prog, "#ffca28")]
        total = max(tot, 1)
        bar = "".join(
            '<div class="seg" style="width:%.3f%%;background:%s;" title="%s: %s"></div>' % (
                v / total * 100, color, label, fmt_en(v))
            for label, v, color in segs if v > 0
        )
        lab_rows.append(
            '<div class="lab-row"><div class="lab-row-lbl">%s <span class="lab-row-n">(%s tests)</span></div>'
            '<div class="stackbar">%s</div></div>' % (name, fmt_en(tot), bar))
    chart_lab_rows = "\n".join(lab_rows)
    chart_lab_legend = '<div class="legend">' + "".join(
        '<div class="legend-item"><span class="dot" style="background:%s;"></span>%s</div>' % (c, l)
        for l, c in (("Pass", "#00e676"), ("Fail", "#ff4757"), ("In Progress", "#ffca28"))
    ) + "</div>"

    return """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Sakarya 3LPP Inspection - Weekly Summary</title>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@300;400;500&display=swap" rel="stylesheet">
<style>
:root{{--bg:#060d1a;--s1:#0c1929;--s2:#102035;--border:rgba(0,168,255,0.12);--blue:#00a8ff;--teal:#00e5c8;--purple:#7c4dff;--green:#00e676;--amber:#ffca28;--red:#ff4757;--text:#ddeeff;--muted:#4a7090;--r:10px;}}
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Outfit',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;padding:0 0 60px;}}
header{{display:flex;justify-content:space-between;align-items:center;padding:20px 52px;border-bottom:1px solid var(--border);flex-wrap:wrap;gap:10px;}}
.brand-title{{font-size:18px;font-weight:700;letter-spacing:1px;background:linear-gradient(90deg,var(--blue),var(--teal));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}}
.brand-sub{{font-family:'JetBrains Mono',monospace;font-size:9px;letter-spacing:3px;color:var(--muted);text-transform:uppercase;margin-top:3px;}}
.proj-info{{font-family:'JetBrains Mono',monospace;font-size:9px;color:var(--muted);margin-top:6px;}}
.stat-row{{display:flex;flex-wrap:wrap;gap:28px;}}
.stat-item .stat-label{{font-family:'JetBrains Mono',monospace;font-size:9px;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);margin-bottom:6px;}}
.stat-item .stat-val{{font-size:28px;font-weight:800;color:var(--text);}}
.chart-note{{font-size:10px;color:var(--muted);margin-top:14px;font-style:italic;}}
.rm-table{{width:100%;border-collapse:collapse;font-size:12px;}}
.rm-table th{{text-align:left;font-family:'JetBrains Mono',monospace;font-size:9px;letter-spacing:1px;text-transform:uppercase;color:var(--muted);padding:6px 10px;border-bottom:1px solid var(--border);}}
.rm-table td{{padding:9px 10px;border-bottom:1px solid var(--border);color:var(--text);}}
.rm-table tr:last-child td{{border-bottom:none;}}
.rm-badge{{display:inline-block;margin-left:8px;padding:2px 8px;border-radius:10px;background:rgba(255,202,40,.15);color:var(--amber);font-size:8px;font-family:'JetBrains Mono',monospace;letter-spacing:.5px;white-space:nowrap;}}
.hdr-date{{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--muted);}}
.hdr-date b{{color:var(--text);}}
a.back{{color:var(--blue);font-family:'JetBrains Mono',monospace;font-size:10px;letter-spacing:1px;text-decoration:none;}}
a.back:hover{{text-decoration:underline;}}
main{{padding:32px 52px;max-width:1200px;margin:0 auto;}}
.sl{{display:flex;align-items:center;gap:14px;font-family:'JetBrains Mono',monospace;font-size:9px;letter-spacing:4px;text-transform:uppercase;color:var(--blue);margin:28px 0 16px;}}
.sl::before{{content:'';width:20px;height:2px;background:linear-gradient(90deg,var(--blue),var(--teal));border-radius:1px;}}
.sl::after{{content:'';flex:1;height:1px;background:var(--border);}}
.kpi-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;}}
.kpi{{background:var(--s1);border:1px solid var(--border);border-radius:var(--r);padding:20px 24px;position:relative;overflow:hidden;transition:transform .2s,border-color .2s,box-shadow .2s;}}
.kpi::after,.chart-card::after{{content:'';position:absolute;inset:0;border-radius:inherit;background:radial-gradient(220px circle at var(--mx,50%) var(--my,50%),color-mix(in srgb,var(--blue) 16%,transparent),transparent 65%);opacity:0;transition:opacity .25s;pointer-events:none;}}
.kpi:hover::after,.chart-card:hover::after{{opacity:1;}}
.kpi:hover{{transform:translateY(-3px);border-color:rgba(0,168,255,.4);box-shadow:0 8px 24px rgba(0,168,255,.08);}}
.kpi:hover .kpi-bar{{opacity:1;}}
.kpi-label{{font-family:'JetBrains Mono',monospace;font-size:9px;letter-spacing:2px;text-transform:uppercase;color:var(--muted);margin-bottom:10px;}}
.kpi-val{{font-size:32px;font-weight:800;line-height:1.15;margin-bottom:6px;white-space:nowrap;background:linear-gradient(135deg,var(--blue),var(--teal));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}}
.kpi-sub{{font-size:11px;color:var(--muted);}}
.kpi-bar{{position:absolute;bottom:0;left:0;right:0;height:3px;opacity:.85;transition:opacity .2s;}}
.kpi-hero{{display:flex;align-items:center;gap:26px;margin-bottom:14px;padding:20px 28px;}}
.ring-wrap{{position:relative;width:90px;height:90px;flex-shrink:0;}}
.progress-ring{{transform:rotate(-90deg);}}
.ring-bg{{fill:none;stroke:var(--border);stroke-width:7;}}
.ring-fill{{fill:none;stroke:url(#ringGrad);stroke-width:7;stroke-linecap:round;stroke-dasharray:{circumference}px;}}
.ring-pct{{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:15px;font-weight:800;color:var(--text);}}
.kpi-hero-text{{min-width:0;}}
.kpi-hero-date{{font-size:26px;font-weight:800;color:var(--text);margin:4px 0 6px;white-space:nowrap;}}
@media (prefers-reduced-motion: no-preference) {{
  .ring-fill{{animation:ringFill 1.4s cubic-bezier(.4,0,.2,1) forwards;}}
}}
@keyframes ringFill{{from{{stroke-dashoffset:{circumference}px;}}to{{stroke-dashoffset:{ring_offset}px;}}}}
.card{{background:var(--s1);border:1px solid var(--border);border-radius:var(--r);padding:24px;margin-bottom:14px;}}
ul.bullets{{list-style:none;display:flex;flex-direction:column;gap:10px;}}
ul.bullets li{{font-size:13px;line-height:1.5;padding-left:16px;position:relative;color:var(--text);}}
ul.bullets li::before{{content:'';position:absolute;left:0;top:7px;width:6px;height:6px;border-radius:50%;background:var(--teal);}}
.chart-grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;}}
.chart-card{{background:var(--s1);border:1px solid var(--border);border-radius:var(--r);padding:22px 24px;position:relative;overflow:hidden;transition:border-color .2s,box-shadow .2s;}}
.chart-card:hover{{border-color:rgba(0,168,255,.25);box-shadow:0 4px 20px rgba(0,168,255,.05);}}
.chart-card.wide{{grid-column:1 / -1;}}
.chart-title{{font-size:13px;font-weight:600;color:var(--text);margin-bottom:16px;}}
.stackbar{{display:flex;height:10px;border-radius:4px;overflow:hidden;background:rgba(255,255,255,.04);gap:2px;}}
.seg{{height:100%;min-width:2px;}}
.legend{{display:flex;flex-wrap:wrap;gap:16px;margin-top:14px;}}
.legend-item{{display:flex;align-items:center;gap:7px;font-size:11px;color:var(--muted);}}
.legend-item b{{color:var(--text);font-weight:600;margin-left:2px;}}
.legend-item .lg-pct{{color:var(--muted);font-family:'JetBrains Mono',monospace;font-size:10px;}}
.dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0;}}
.lab-row{{margin-bottom:16px;}}
.lab-row:last-child{{margin-bottom:0;}}
.lab-row-lbl{{font-size:12px;color:var(--text);margin-bottom:7px;}}
.lab-row-n{{color:var(--muted);font-family:'JetBrains Mono',monospace;font-size:10px;}}
.completion-row{{display:flex;align-items:center;gap:28px;}}
.completion-pct{{font-size:44px;font-weight:800;line-height:1;flex-shrink:0;background:linear-gradient(135deg,var(--blue),var(--teal));-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}}
.completion-bar-wrap{{flex:1;}}
.completion-bar-wrap .stackbar{{height:14px;border-radius:7px;}}
@media (max-width:640px){{.completion-row{{flex-direction:column;align-items:flex-start;gap:16px;}}}}
.print-btn{{display:flex;align-items:center;gap:8px;padding:7px 18px;background:linear-gradient(90deg,var(--blue),var(--teal));border:none;border-radius:6px;font-family:'JetBrains Mono',monospace;font-size:9px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:#060d1a;cursor:pointer;transition:opacity .2s;}}
.print-btn:hover{{opacity:.85;}}
footer{{border-top:1px solid var(--border);background:rgba(6,13,26,.95);padding:28px 52px;margin-top:40px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:24px;}}
.ft-brand{{display:flex;align-items:center;gap:18px;}}
.ft-div{{width:1px;height:40px;background:var(--border);}}
.ft-co{{display:flex;flex-direction:column;gap:3px;}}
.ft-name{{font-size:12px;font-weight:600;color:var(--text);}}
.ft-tag{{font-family:'JetBrains Mono',monospace;font-size:8px;letter-spacing:2px;color:var(--muted);text-transform:uppercase;}}
.ft-contacts{{display:flex;gap:36px;flex-wrap:wrap;}}
.ft-col{{display:flex;flex-direction:column;gap:4px;}}
.ft-col-lbl{{font-family:'JetBrains Mono',monospace;font-size:7px;letter-spacing:2px;text-transform:uppercase;color:var(--blue);margin-bottom:2px;}}
.ft-col span,.ft-col a{{font-size:11px;color:var(--muted);text-decoration:none;}}
.ft-col a:hover{{color:var(--blue);}}
.ft-stamp{{font-family:'JetBrains Mono',monospace;font-size:8px;color:rgba(74,112,144,.3);letter-spacing:1px;text-align:right;align-self:flex-end;}}
@media print{{
  *{{-webkit-print-color-adjust:exact!important;print-color-adjust:exact!important;box-shadow:none!important;}}
  @page{{size:A4;margin:14mm;}}
  .no-print{{display:none!important;}}
  body{{background:#fff;color:#1a2233;padding:0;}}
  header{{border-bottom:1px solid #ccc;padding:0 0 12px;}}
  .brand-title{{background:none;-webkit-text-fill-color:#0a3d62;color:#0a3d62;}}
  .brand-sub,.hdr-date,.proj-info{{color:#555;}}
  .hdr-date b{{color:#1a2233;}}
  main{{padding:16px 0 0;max-width:none;}}
  .sl{{color:#0a3d62;margin:18px 0 10px;}}
  .sl::after{{background:#ccc;}}
  .kpi,.chart-card,.card{{background:#f7f9fb;border:1px solid #ddd;}}
  .kpi::after,.chart-card::after{{display:none;}}
  .kpi-val{{background:none;-webkit-text-fill-color:#0a3d62;color:#0a3d62;}}
  .kpi-hero-date{{color:#0a3d62;}}
  .ring-pct{{color:#1a2233;}}
  .ring-bg{{stroke:#ddd;}}
  .kpi-sub,.legend-item,.lab-row-n,.lg-pct,.stat-item .stat-label,.chart-note{{color:#555;}}
  .legend-item b,.lab-row-lbl,.stat-item .stat-val{{color:#1a2233;}}
  .rm-table td{{color:#1a2233;border-bottom:1px solid #ddd;}}
  .rm-table th{{color:#555;border-bottom:1px solid #ddd;}}
  .rm-badge{{background:#fef3c7;color:#92400e;}}
  footer{{background:#fff;border-top:1px solid #ccc;}}
  .ft-name{{color:#1a2233;}}
  .ft-tag,.ft-col span,.ft-col a{{color:#555;}}
  .ft-col-lbl{{color:#0a3d62;}}
  .ft-stamp{{color:#999;}}
  ul.bullets li{{color:#1a2233;}}
  .chart-grid{{grid-template-columns:1fr 1fr;}}
  .chart-card.wide{{grid-column:1 / -1;}}
  .completion-pct{{background:none;-webkit-text-fill-color:#0a3d62;color:#0a3d62;}}
  .kpi,.chart-card,.card{{page-break-inside:avoid;}}
}}
/* Tema chiaro: applicato in automatico se scelto nel dashboard principale
   (localStorage "sak_theme", stesso meccanismo di index.html) - nessun
   pulsante THEME qui, la pagina eredita e basta. */
body.light{{--border:rgba(0,100,200,0.15);--text:#0f172a;--muted:#64748b;}}
body.light header{{background:rgba(240,244,248,0.9);}}
body.light .brand-title,body.light .completion-pct{{background:none;-webkit-text-fill-color:#0050a0;color:#0050a0;}}
body.light .kpi-val{{background:none;-webkit-text-fill-color:unset;color:#0050a0;}}
body.light .stackbar{{background:rgba(0,0,0,.05);}}
</style>
</head>
<body>
<script>
// Eredita TEMA e TWEAKS scelti nel dashboard principale (index.html) -
// stesse chiavi localStorage, nessun pulsante THEME/TWEAKS qui: la pagina
// eredita e basta.
(function(){{
  // ---- Tema (sfondo + chiaro/scuro), chiave "sak_theme" ----
  try {{
    var saved = localStorage.getItem("sak_theme");
    if (saved) {{
      var t = JSON.parse(saved);
      var root = document.documentElement;
      if (t.bg) root.style.setProperty("--bg", t.bg);
      if (t.s1) root.style.setProperty("--s1", t.s1);
      if (t.s2) root.style.setProperty("--s2", t.s2);
      if (t.isLight) document.body.classList.add("light");
    }}
  }} catch(e) {{}}

  // ---- Tweaks (palette accento / densita' / movimento), chiave "sakarya-tweaks-v1" ----
  var ACCENTS = {{
    ocean:  {{ blue:"#00a8ff", teal:"#00e5c8", purple:"#7c4dff" }},
    ember:  {{ blue:"#ffb020", teal:"#ff6b35", purple:"#ff4757" }},
    signal: {{ blue:"#8b5cf6", teal:"#22d3ee", purple:"#ec4899" }}
  }};
  var DENSITY_CSS = {{
    compact: "main{{padding:20px 32px !important;}}.kpi-grid{{gap:10px !important;}}.kpi{{padding:14px 16px !important;}}.kpi-val{{font-size:26px !important;}}.chart-grid{{gap:10px !important;}}.chart-card{{padding:16px 18px !important;}}.card{{padding:16px !important;}}.sl{{margin:20px 0 12px !important;}}",
    standard: "",
    spacious: "main{{padding:48px 68px !important;}}.kpi-grid{{gap:20px !important;}}.kpi{{padding:28px 32px !important;}}.kpi-val{{font-size:42px !important;}}.chart-grid{{gap:20px !important;}}.chart-card{{padding:30px 32px !important;}}.card{{padding:32px !important;}}.sl{{margin:36px 0 20px !important;}}"
  }};
  var MOTION_CSS = {{
    calm: "*{{transition-duration:.1s !important;}}.kpi:hover{{transform:none !important;}}.chart-card:hover{{box-shadow:none !important;}}",
    standard: "",
    expressive: ".kpi:hover{{transform:translateY(-5px) scale(1.01) !important;box-shadow:0 14px 30px rgba(0,168,255,.15) !important;}}.chart-card:hover{{transform:translateY(-3px) !important;box-shadow:0 10px 26px rgba(0,168,255,.15) !important;}}"
  }};
  function styleTag(id) {{
    var el = document.getElementById(id);
    if (!el) {{ el = document.createElement("style"); el.id = id; document.head.appendChild(el); }}
    return el;
  }}
  try {{
    var tw = JSON.parse(localStorage.getItem("sakarya-tweaks-v1") || "{{}}");
    var accent = ACCENTS[tw.accent] || ACCENTS.ocean;
    var root2 = document.documentElement;
    root2.style.setProperty("--blue", accent.blue);
    root2.style.setProperty("--teal", accent.teal);
    root2.style.setProperty("--purple", accent.purple);
    styleTag("tweak-density-style").textContent = DENSITY_CSS[tw.density] || "";
    styleTag("tweak-motion-style").textContent = MOTION_CSS[tw.motion] || "";
  }} catch(e) {{}}
}})();
</script>
<header>
  <div>
    <div class="brand-title">Sakarya Gas Field Development - 3LPP Inspection</div>
    <div class="brand-sub">Weekly Summary - Coating Plant #6 - Tenaris Confab</div>
    <div class="proj-info">QCP No: {qcp_no} &nbsp;&middot;&nbsp; {od_wt}</div>
  </div>
  <div style="display:flex;align-items:center;gap:20px;">
    <a class="back no-print" href="index.html">&larr; Overall Status Dashboard</a>
    <div class="hdr-date">Report Date &nbsp;<b>{report_date}</b></div>
    <button class="print-btn no-print" onclick="window.print()"><svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5"><polyline points="6 9 6 2 18 2 18 9"/><path d="M6 18H4a2 2 0 01-2-2v-5a2 2 0 012-2h16a2 2 0 012 2v5a2 2 0 01-2 2h-2"/><rect x="6" y="14" width="12" height="8"/></svg>Print Report</button>
  </div>
</header>

<main>
  <div class="sl">Key Performance Indicators</div>
  <div class="kpi kpi-hero">
    <div class="ring-wrap">
      <svg class="progress-ring" viewBox="0 0 90 90" width="90" height="90">
        <defs><linearGradient id="ringGrad" x1="0%" y1="0%" x2="100%" y2="100%">
          <stop offset="0%" stop-color="var(--blue)"/><stop offset="100%" stop-color="var(--teal)"/>
        </linearGradient></defs>
        <circle class="ring-bg" cx="45" cy="45" r="38"></circle>
        <circle class="ring-fill" cx="45" cy="45" r="38" style="stroke-dashoffset:{ring_offset}px;"></circle>
      </svg>
      <div class="ring-pct">{pct_complete}</div>
    </div>
    <div class="kpi-hero-text">
      <div class="kpi-label">Est. Completion</div>
      <div class="kpi-hero-date">{production_estimation}</div>
      <div class="kpi-sub">{work_days_to_finish} work days remaining &middot; {pct_complete} of order complete</div>
    </div>
  </div>
  <div class="kpi-grid">
    <div class="kpi"><div class="kpi-label">Pipes to Finish</div><div class="kpi-val">{tot_pipes_to_finish}</div><div class="kpi-sub">of {tot_item_qty} total item qty &middot; {work_days_to_finish} work days</div><div class="kpi-bar" style="background:linear-gradient(90deg,var(--blue),var(--teal));"></div></div>
    <div class="kpi"><div class="kpi-label">Total Coated (VDI+Final)</div><div class="kpi-val">{coated_totale}</div><div class="kpi-sub">VDI: {coated_vdi} | Final: {coated_fin}</div><div class="kpi-bar" style="background:linear-gradient(90deg,var(--blue),var(--teal));"></div></div>
    <div class="kpi"><div class="kpi-label">Approved - Final Inspection</div><div class="kpi-val">{app_fin}</div><div class="kpi-sub">{pct_app_fin} of final coated</div><div class="kpi-bar" style="background:linear-gradient(90deg,var(--blue),var(--teal));"></div></div>
    <div class="kpi"><div class="kpi-label">Not Approved / Quarantine</div><div class="kpi-val" style="background:none;-webkit-text-fill-color:{col_na};color:{col_na};">{on_hold_count}</div><div class="kpi-sub">{pct_na} of total coated</div><div class="kpi-bar" style="background:{col_na};"></div></div>
    <div class="kpi"><div class="kpi-label">Lab Tests Failed (PPT+Prod)</div><div class="kpi-val" style="background:none;-webkit-text-fill-color:{col_fail};color:{col_fail};">{fail_lab_totale}</div><div class="kpi-sub">{pct_fail} of {tot_lab} tests performed</div><div class="kpi-bar" style="background:{col_fail};"></div></div>
    <div class="kpi"><div class="kpi-label">Stripping Control Open</div><div class="kpi-val">{strip_count}</div><div class="kpi-sub">non-conformances recorded</div><div class="kpi-bar" style="background:linear-gradient(90deg,var(--blue),var(--teal));"></div></div>
  </div>

  <div class="sl">3LPP Raw Material</div>
  <div class="card">
    <table class="rm-table">
      <thead><tr><th></th><th>Manufacturer</th><th>Trade Name</th><th>Batch</th></tr></thead>
      <tbody>{rm_table_rows}</tbody>
    </table>
    <div class="chart-note">Materials in use as of last logged shift ({rm_date}). Source: "Raw Material" sheet, columns C-Q.</div>
  </div>

  <div class="sl">Analysis &amp; Observations</div>
  <div class="card"><ul class="bullets">{bullets_html}</ul></div>

  <div class="sl">Charts</div>
  <div class="chart-grid">
    <div class="chart-card wide">
      <div class="chart-title">Order Completion</div>
      <div class="completion-row">
        <div class="completion-pct">{pct_complete}</div>
        <div class="completion-bar-wrap">{chart_completion}</div>
      </div>
    </div>
    <div class="chart-card">
      <div class="chart-title">Total Coated by Station</div>
      {chart_coated}
    </div>
    <div class="chart-card">
      <div class="chart-title">Not Approved / Quarantine by Disposition</div>
      {chart_quarantine}
    </div>
    <div class="chart-card">
      <div class="chart-title">On Hold</div>
      <div class="stat-row">{oh_summary_html}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">After Action Outcome</div>
      {chart_action}
    </div>
    <div class="chart-card wide">
      <div class="chart-title">Production Throughput by Station</div>
      {chart_station_rows}
      {chart_station_legend}
    </div>
    <div class="chart-card wide">
      <div class="chart-title">Lab Test Results by Type</div>
      {chart_lab_rows}
      {chart_lab_legend}
    </div>
  </div>

</main>
<footer>
  <div class="ft-brand">
    <img src="{iqs_logo}" alt="IQS" style="height:36px;width:auto;">
    <div class="ft-div"></div>
    <div class="ft-co"><div class="ft-name">IQS S.r.l. Worldwide Inspection Company</div><div class="ft-tag">Global Quality Assurance</div></div>
  </div>
  <div class="ft-contacts">
    <div class="ft-col"><div class="ft-col-lbl">Contact</div><span>+39 02 49532180 / 2181</span><a href="https://www.iqs-ww.com">www.iqs-ww.com</a></div>
    <div class="ft-col"><div class="ft-col-lbl">Registered Office</div><span>Via Trieste 16/A</span><span>20097 S. Donato Mil.se (MI)</span></div>
    <div class="ft-col"><div class="ft-col-lbl">Head Quarter</div><span>Via Europa 42</span><span>20097 S. Donato Mil.se (MI)</span></div>
  </div>
  <div class="ft-stamp">IQS Confidential &middot; Report {report_date}</div>
</footer>
<script>
// Cursor-following glow on cards (same effect as the main dashboard)
document.addEventListener("mousemove", (e) => {{
  const el = e.target.closest(".kpi,.chart-card");
  if (!el) return;
  const r = el.getBoundingClientRect();
  el.style.setProperty("--mx", (e.clientX - r.left) + "px");
  el.style.setProperty("--my", (e.clientY - r.top) + "px");
}}, {{ passive: true }});
</script>
</body>
</html>
""".format(
        report_date=report_date, iqs_logo=IQS_LOGO_B64,
        qcp_no=kpi["qcp_no"] or "n/a", od_wt=kpi["od_wt"] or "n/a",
        production_estimation=fmt_date(kpi["production_estimation"]),
        work_days_to_finish=fmt_en(kpi["work_days_to_finish"]),
        tot_pipes_to_finish=fmt_en(kpi["tot_pipes_to_finish"]), tot_item_qty=fmt_en(kpi["tot_item_qty"]),
        coated_totale=fmt_en(kpi["coated_totale"]), coated_vdi=fmt_en(kpi["coated_vdi"]), coated_fin=fmt_en(kpi["coated_fin"]),
        app_fin=fmt_en(kpi["app_fin"]), pct_app_fin=fmt_pct(kpi["pct_app_fin"]),
        on_hold_count=fmt_en(kpi["on_hold_count"]), pct_na=fmt_pct(kpi["pct_na"]), col_na=col_na,
        fail_lab_totale=fmt_en(kpi["fail_lab_totale"]), pct_fail=fmt_pct(kpi["pct_fail"]), tot_lab=fmt_en(kpi["tot_lab"]), col_fail=col_fail,
        strip_count=fmt_en(kpi["strip_count"]),
        bullets_html=bullets_html,
        pct_complete=fmt_pct(kpi["pct_complete"]), chart_completion=chart_completion,
        circumference="%.2f" % circumference, ring_offset="%.2f" % ring_offset,
        chart_coated=chart_coated, chart_quarantine=chart_quarantine,
        oh_summary_html=oh_summary_html, chart_action=chart_action,
        rm_table_rows=rm_table_rows, rm_date=rm_date,
        chart_station_rows=chart_station_rows, chart_station_legend=chart_station_legend,
        chart_lab_rows=chart_lab_rows, chart_lab_legend=chart_lab_legend,
    )


def main():
    excel = trova_file_excel()
    if not excel:
        log("[ERRORE] Nessun file '%s' trovato in questa cartella." % PATTERN)
        sys.exit(1)

    report_date = data_da_nome_file(excel)
    log("=" * 60)
    log("  SAKARYA 3LPP INSPECTION - generazione sunto")
    log("=" * 60)
    log("File Excel  : %s" % os.path.basename(excel))
    log("Report Date : %s" % report_date)

    try:
        kpi = estrai_kpi(excel)
    except Exception as e:
        log("[ERRORE] %s" % e)
        sys.exit(1)

    html = genera_html(kpi, report_date)
    out_path = os.path.join(HERE, "3lpp.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)

    log("-" * 60)
    log("KPI estratti:")
    log("  Coated totale        : %s (VDI %s / Finale %s)" % (kpi["coated_totale"], kpi["coated_vdi"], kpi["coated_fin"]))
    log("  Approvate Finale     : %s (%.1f%%)" % (kpi["app_fin"], kpi["pct_app_fin"] * 100))
    log("  Non Approvate/Quar.  : %s (%.1f%%)" % (kpi["on_hold_count"], kpi["pct_na"] * 100))
    log("  Lab Tests Falliti    : %s / %s (%.1f%%)" % (kpi["fail_lab_totale"], kpi["tot_lab"], kpi["pct_fail"] * 100))
    log("  Stripping Aperti     : %s" % kpi["strip_count"])
    log("-" * 60)
    log("OK - pagina generata: %s" % out_path)


if __name__ == "__main__":
    main()
