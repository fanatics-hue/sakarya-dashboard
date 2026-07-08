# -*- coding: utf-8 -*-
"""
genera_dati.py — Sakarya QC Dashboard
Legge il file Excel "Sakarya Inspection Overall Status as of *.xlsx" piu' recente
e genera 3 file CSV in ./data/ che il dashboard (index.html) legge direttamente:
    data/summary.csv        -> KPI + ITP Steps  (scheda Dashboard dell'Excel)
    data/defectsWeld.csv    -> difetti saldatura (scheda Repair)
    data/defectsFinal.csv   -> difetti final     (scheda Rejection)

NON serve aprire o modificare nulla a mano: doppio-clic su AGGIORNA_DASHBOARD.bat.
"""
import csv
import glob
import os
import re
import sys
from datetime import datetime

import openpyxl

# ======================================================================
#  CONFIG — l'unico valore che NON e' nell'Excel di status.
#  Modificalo qui solo se cambia (raramente).
# ======================================================================
PO_QTY = 15462          # Purchase Order (PO) Qty — quantita' totale ordine
# NB: "Pipes Rejected" NON e' un config: viene calcolato come numero di
#     righe della tabella Rejection (foglio Rejection -> defectsFinal),
#     replicando la logica del vecchio dashboard live.
# ======================================================================

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")
PATTERN = "Sakarya Inspection Overall Status as of *.xlsx"


def log(msg):
    print(msg, flush=True)


def trovan_file_excel():
    """Trova il file di status piu' recente nella cartella."""
    candidati = [f for f in glob.glob(os.path.join(HERE, PATTERN))
                 if not os.path.basename(f).startswith("~$")]
    if not candidati:
        log("[ERRORE] Nessun file '%s' trovato nella cartella." % PATTERN)
        sys.exit(1)

    def chiave_data(path):
        # piu' recente per data "as of MM.DD.YYYY" nel nome file: non ci si
        # affida al mtime, che puo' essere alterato da copie/sync e far
        # scegliere per sbaglio un file di una settimana vecchia.
        m = re.search(r"as of (\d{1,2})\.(\d{1,2})\.(\d{4})", os.path.basename(path))
        if m:
            mm, dd, yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return (yyyy, mm, dd)
        return (0, 0, 0)

    f = max(candidati, key=lambda p: (chiave_data(p), os.path.getmtime(p)))
    return f


def data_da_nome_file(path):
    """Estrae 'as of MM.DD.YYYY' dal nome file -> stringa DD/MM/YYYY."""
    m = re.search(r"as of (\d{1,2})\.(\d{1,2})\.(\d{4})", os.path.basename(path))
    if m:
        mm, dd, yyyy = m.group(1), m.group(2), m.group(3)
        return "%02d/%02d/%s" % (int(dd), int(mm), yyyy)
    return datetime.now().strftime("%d/%m/%Y")


def fmt_data(v):
    """Converte un valore data (datetime o stringa) in DD/MM/YYYY."""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    s = str(v or "").strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return "%s/%s/%s" % (m.group(3), m.group(2), m.group(1))
    return s.split(" ")[0]


def num(v, default=0):
    """Estrae un numero da una cella (gestisce None, stringhe, errori #DIV/0)."""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip().replace(",", ".")
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group(0)) if m else default


def righe_foglio(ws, max_col=12):
    """Restituisce tutte le righe del foglio come liste di celle (valori)."""
    return [list(r) for r in ws.iter_rows(values_only=True)]


# ======================================================================
#  ITP Steps dal registro per-tubo (es. foglio "PO5021701")
#
#  Il foglio "Dashboard" ha una tabella ITP Steps mantenuta a mano/con
#  formule Excel che puo' restare ferma a un numero vecchio di tubi
#  (capitato: fermo a 170 mentre il registro reale ne aveva gia' 2602).
#  Il registro per-tubo (una riga per tubo, una colonna per stazione,
#  cella piena = stazione superata) e' la fonte diretta e sempre
#  aggiornata: qui si CONTANO le celle piene per colonna invece di
#  fidarsi del riepilogo.
#
#  Il nome del foglio cambia (es. col numero di lotto/PO Tenaris), quindi
#  NON si cerca per nome ma per intestazione (riga 2: "Plate Entrance" +
#  "Pipe Forming" sono la firma di questo registro).
# ======================================================================
ITP_STEPS_CANONICI = [
    # (Seq, Nome step, Sezione, colonna nel registro per-tubo o None)
    (1,  "Incoming Plate",       "Forming",             "Plate Entrance"),
    (2,  "Pipe Forming",         "Forming",             "Pipe Forming"),
    (3,  "Tack Welding",         "Forming",             "Tack Welding"),
    (4,  "ID SAWL",              "Forming",             "ID SAWL"),
    (5,  "OD SAWL",              "Forming",             "OD SAWL"),
    (6,  "AUT1",                 "Welding & NDT",       "AUT1 in House"),
    (7,  "Fluoroscopic",         "Welding & NDT",       "Fluoroscopic/TOFD In House"),
    (8,  "Welding Repair",       "Welding & NDT",       "Repair"),
    (9,  "MCE1",                 "Mechanical",          "MCE 1"),
    (10, "MCE2",                 "Mechanical",          "MCE 2"),
    (11, "MCE Inspection",       "Mechanical",          None),  # segue MCE1 (nessuna colonna propria nel registro)
    (12, "Straightness Re-work", "Mechanical",          "Straighness Re-work"),
    (13, "Hydro Test",           "Testing",             "Hydrostatic Test"),
    (14, "AUT2",                 "Testing",             None),  # step non tracciato nel registro
    (15, "Phase Arr.",           "Testing",             "Phase Array"),
    (16, "MUT",                  "Testing",             "MUT"),
    (17, "End UT",               "Testing",             "Pipe End UT"),
    (18, "X-Ray",                "Testing",             "X-Ray"),
    (19, "OoR Pipe Body",        "Dimensional & Final", "OoR Pipe Body"),
    (20, "Dimensional 1",        "Dimensional & Final", "Dimensional 1"),
    (21, "Dimensional 2",        "Dimensional & Final", "Dimensional 2"),
    (22, "MPI",                  "Dimensional & Final", "MPI"),
    (23, "UT DHC",               "Dimensional & Final", "UT DHC"),
    (24, "Final Mark & Weigh",   "Dimensional & Final", "Final Marking and Weighing"),
]


def norm_pipe(pipe):
    """Normalizza un Pipe N° (es. '26.4.002086' o '26.4.2086') in modo che i
    due formati usati nei diversi fogli (con/senza zeri iniziali) coincidano."""
    if not pipe:
        return None
    parts = str(pipe).strip().split(".")
    if len(parts) >= 3:
        try:
            return "%d.%d.%d" % (int(parts[0]), int(parts[1]), int(parts[2]))
        except ValueError:
            return str(pipe).strip()
    return str(pipe).strip()


def trova_registro_per_tubo(wb):
    """Trova il foglio registro per-tubo cercando la firma nell'intestazione
    (non per nome: il nome del foglio cambia col lotto/PO Tenaris)."""
    for name in wb.sheetnames:
        rows = righe_foglio(wb[name])
        for r in rows[:5]:
            testo = {str(c).strip() for c in r if c}
            if "Plate Entrance" in testo and "Pipe Forming" in testo:
                return name, rows
    return None, None


def estrai_itp_da_registro(rows):
    """Conta le celle piene per colonna nel registro per-tubo e mappa
    sui 24 step canonici. Ritorna (lista_itp, incoming_count)."""
    header = rows[1]
    dati = rows[2:]
    col_idx = {str(h).strip(): i for i, h in enumerate(header) if h}
    conteggi = {}
    for nome_col, i in col_idx.items():
        conteggi[nome_col] = sum(1 for r in dati if i < len(r) and r[i] not in (None, ""))

    incoming = conteggi.get("Plate Entrance", len(dati))
    itp = []
    for seq, name, group, col in ITP_STEPS_CANONICI:
        if col is not None:
            count = conteggi.get(col, 0)
        elif name == "MCE Inspection":
            count = conteggi.get("MCE 1", 0)
        else:
            count = 0
        loss = max(0, incoming - count)
        yld = round(count / incoming * 100, 2) if incoming else 0
        itp.append([seq, name, group, count, yld, loss, loss, ""])
    return itp, incoming


def calcola_accepted(rows):
    """'Pipes Accepted' (etichettato "Ready to Coating" in index.html) = tubi
    che hanno completato la colonna 'UT DHC' nel registro per-tubo (scelta di
    Rino 2026-07-08: e' l'ultimo controllo NDT prima che il tubo sia pronto
    per il coating 3LPP — NON piu' "Final Marking and Weighing", che e' l'ULTIMO
    step dell'intero ITP e rappresenta un traguardo diverso, piu' a valle).
    Ritorna (n_accettati, lunghezza_totale_mm) — lunghezza_totale_mm e' None
    se non si trova una colonna lunghezza."""
    header = rows[1]
    # "Pipe N°" ha un problema di encoding nel file sorgente (il ° arriva
    # come carattere corrotto): si individua per prefisso, escludendo le
    # colonne "Pipe Number"/"Pipe Dimensional" del blocco del lotto corrente.
    idx_pipe = next((i for i, h in enumerate(header)
                      if h and str(h).lower().startswith("pipe n")
                      and "number" not in str(h).lower()
                      and "dimensional" not in str(h).lower()), None)
    idx_final = next((i for i, h in enumerate(header)
                       if h == "UT DHC"), None)
    dati = rows[2:]
    if idx_pipe is None or idx_final is None:
        return None, None

    # colonna lunghezza del REGISTRO CUMULATIVO (dopo idx_pipe): nel file
    # con doppio blocco esiste anche una "Pipe length" per il solo lotto
    # corrente, prima di idx_pipe, che va ignorata.
    idx_len = next((i for i, h in enumerate(header)
                     if i > idx_pipe and h and "length" in str(h).lower()), None)

    finished = set()
    lunghezza_totale = 0
    for r in dati:
        if idx_pipe < len(r) and idx_final < len(r) and r[idx_pipe] and r[idx_final]:
            pipe = norm_pipe(r[idx_pipe])
            if pipe in finished:
                continue
            finished.add(pipe)
            if idx_len is not None and idx_len < len(r):
                lunghezza_totale += num(r[idx_len], 0)
    return len(finished), (lunghezza_totale if idx_len is not None else None)


# ----------------------------------------------------------------------
#  ESTRAZIONE
# ----------------------------------------------------------------------
def estrai_summary(wb, report_date, rejected):
    """Dal foglio 'Dashboard': KPI + 24 ITP Steps.
    `rejected` = n. righe tabella Rejection (calcolato a parte)."""
    ws = wb["Dashboard"]
    rows = righe_foglio(ws)

    def cella_dopo_label(label_sub):
        """Trova la prima riga con una cella che contiene label_sub e
        ritorna il primo valore numerico successivo nella stessa riga."""
        label_sub = label_sub.lower()
        for r in rows:
            testo = [("" if c is None else str(c)) for c in r]
            for j, c in enumerate(testo):
                if label_sub in c.lower():
                    for c2 in r[j + 1:]:
                        if isinstance(c2, (int, float)):
                            return c2
        return None

    # --- ITP Steps: dal registro per-tubo (conteggio diretto, sempre
    # aggiornato) invece che dal riepilogo del foglio Dashboard, che puo'
    # restare fermo a un numero vecchio di tubi (vedi commento sopra
    # trova_registro_per_tubo).
    foglio_reg, righe_reg = trova_registro_per_tubo(wb)
    if foglio_reg:
        itp, incoming = estrai_itp_da_registro(righe_reg)
        repair_count = next((r[3] for r in itp if r[1] == "Welding Repair"), 0)

        # Accepted = tubi che hanno finito il registro (foglio Dashboard:
        # stesso problema di staleness di Incoming Plates). lunghezza_mm =
        # somma "Pipe length" dei tubi accettati (in mm, come nel registro).
        accepted, lunghezza_mm = calcola_accepted(righe_reg)
        if accepted is None:
            accepted = int(num(cella_dopo_label("accept"), 0))
        passrate = round(accepted / incoming * 100, 2) if incoming else 0
    else:
        # Fallback: nessun registro per-tubo trovato, si torna al vecchio
        # riepilogo del foglio Dashboard (compatibilita' con Excel piu' vecchi).
        accepted = int(num(cella_dopo_label("accept"), 0))
        lunghezza_mm = None
        passrate = num(cella_dopo_label("overall status"), 0)
        if passrate <= 1.5:      # memorizzato come frazione (0.8765)
            passrate = passrate * 100
        passrate = round(passrate, 2)

        total = cella_dopo_label("total pipes")
        if total is None:
            total = cella_dopo_label("input")
        incoming = int(num(total, 0))

        itp = []
        hdr_idx = -1
        hdr_off = 0
        for i, r in enumerate(rows):
            for j, c in enumerate(r):
                if c and "seq" in str(c).lower():
                    hdr_idx = i
                    hdr_off = j
                    break
            if hdr_idx >= 0:
                break

        repair_count = 0
        if hdr_idx >= 0:
            for r in rows[hdr_idx + 1:]:
                base = r[hdr_off:]
                seq = base[0] if len(base) > 0 else None
                if not isinstance(seq, (int, float)):
                    continue
                name = (base[1] if len(base) > 1 else "") or ""
                group = (base[2] if len(base) > 2 else "") or ""
                count = int(num(base[3] if len(base) > 3 else 0))
                yld = num(base[4] if len(base) > 4 else 0)
                if yld <= 1.5:
                    yld = yld * 100
                loss = int(num(base[5] if len(base) > 5 else 0))
                cumul = int(num(base[6] if len(base) > 6 else 0))
                remarks = ""
                if len(base) > 7 and base[7]:
                    remarks = str(base[7]).strip()
                itp.append([int(seq), str(name).strip(), str(group).strip(),
                            count, round(yld, 2), loss, cumul, remarks])
                if "welding repair" in str(name).strip().lower():
                    repair_count = count

    # ---- scrivi CSV nel formato che il dashboard si aspetta ----
    out = []
    out.append(["SAKARYA GAS FIELD", ""])
    out.append(["KEY PERFORMANCE INDICATORS", ""])
    out.append(["Field", "Value"])
    lunghezza_m = round(lunghezza_mm / 1000, 2) if lunghezza_mm else ""

    out.append(["Purchase Order (PO) Qty", PO_QTY])
    out.append(["Incoming Plates", incoming])
    out.append(["Pipes Accepted", accepted])
    out.append(["Accepted Length (m)", lunghezza_m])
    out.append(["Pipes Rejected", rejected])
    out.append(["Repair / Rework", repair_count])
    out.append(["Overall Pass Rate (%)", passrate])
    out.append(["Report Date", report_date])
    out.append([])
    out.append(["Seq. N°", "ITP Step", "Section", "Count",
                "Yield vs Input (%)", "Loss vs Input", "Cumul. Loss", "Remarks"])
    out.extend(itp)

    return out, dict(incoming=incoming, accepted=accepted, lunghezza_m=lunghezza_m,
                     passrate=passrate, repair=repair_count, itp=len(itp))


def estrai_difetti(wb, sheet_name, con_misure):
    """Dal foglio Repair/Rejection -> tabella difetti standard a 7 colonne."""
    ws = wb[sheet_name]
    rows = righe_foglio(ws)

    # trova header (riga con cella 'DATE') e offset colonna
    hdr_idx, off = -1, 0
    for i, r in enumerate(rows):
        for j, c in enumerate(r):
            if c and str(c).strip().lower().startswith("date"):
                hdr_idx, off = i, j
                break
        if hdr_idx >= 0:
            break
    if hdr_idx < 0:
        return [["Date", "Shift", "Pipe N°", "Defect", "Position", "Extension", "Remarks"]]

    out = [["Date", "Shift", "Pipe N°", "Defect", "Position", "Extension", "Remarks"]]
    for r in rows[hdr_idx + 1:]:
        b = r[off:]
        date = b[0] if len(b) > 0 else None
        pipe = b[2] if len(b) > 2 else None
        if not date or not pipe:
            continue
        shift = str(b[1]).strip() if len(b) > 1 and b[1] is not None else ""
        defect = str(b[3]).strip() if len(b) > 3 and b[3] is not None else ""
        if con_misure:
            pos = int(num(b[4] if len(b) > 4 else 0))
            ext = int(num(b[5] if len(b) > 5 else 0))
            rem = str(b[6]).strip() if len(b) > 6 and b[6] is not None else ""
        else:
            # Rejection: niente colonne Position/Extension
            pos, ext = "", ""
            rem = str(b[4]).strip() if len(b) > 4 and b[4] is not None else ""
        out.append([fmt_data(date), shift, str(pipe).strip(), defect, pos, ext, rem])
    return out


def scrivi_csv(nome, righe):
    path = os.path.join(DATA_DIR, nome)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for r in righe:
            w.writerow(r)
    return path


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    excel = trovan_file_excel()
    report_date = data_da_nome_file(excel)
    log("=" * 60)
    log("  SAKARYA DASHBOARD — generazione dati")
    log("=" * 60)
    log("File Excel  : %s" % os.path.basename(excel))
    log("Report Date : %s" % report_date)

    import warnings
    warnings.simplefilter("ignore")
    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)

    # prima i difetti: "Pipes Rejected" = n. righe della tabella Rejection
    weld = estrai_difetti(wb, "Repair", con_misure=True)
    final = estrai_difetti(wb, "Rejection", con_misure=False)
    rejected = len(final) - 1   # -1 per l'header

    summary, stats = estrai_summary(wb, report_date, rejected)

    scrivi_csv("summary.csv", summary)
    scrivi_csv("defectsWeld.csv", weld)
    scrivi_csv("defectsFinal.csv", final)

    log("-" * 60)
    log("KPI estratti:")
    log("  Incoming Plates  : %s" % stats["incoming"])
    log("  Pipes Accepted   : %s" % stats["accepted"])
    log("  Accepted Length  : %s m" % stats["lunghezza_m"])
    log("  Pass Rate        : %s %%" % stats["passrate"])
    log("  Repair / Rework  : %s" % stats["repair"])
    log("  ITP Steps        : %s righe" % stats["itp"])
    log("  PO Qty (config)  : %s" % PO_QTY)
    log("  Pipes Rejected   : %s  (= righe Rejection)" % rejected)
    log("Difetti Weld (Repair)    : %s righe" % (len(weld) - 1))
    log("Difetti Final (Rejection): %s righe" % (len(final) - 1))
    log("-" * 60)
    log("OK — file scritti in: %s" % DATA_DIR)


if __name__ == "__main__":
    main()
