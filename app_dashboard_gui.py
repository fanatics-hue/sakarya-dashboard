# -*- coding: utf-8 -*-
"""
app_dashboard_gui.py - Sakarya Dashboard (finestra grafica UNICA)

Una sola finestra per aggiornare ENTRAMBE le pagine del sito:
  - "Overall Status"  (file "Sakarya Inspection Overall Status as of ...xlsx")
  - "3LPP Inspection" (file "3LPP Sakarya Inspection Overall Status as of ...xlsm")

Sono due file indipendenti che arrivano separatamente ogni settimana: scegli
quello/i che hai ricevuto (anche uno solo), premi '1. Genera e verifica' per
ciascuno, poi UN SOLO '2. Pubblica su GitHub' pubblica tutto cio' che e'
cambiato (index.html e/o 3lpp.html).

- Se il file scelto non e' gia' nella cartella (o ha un nome diverso), viene
  copiato qui col nome standard "... as of MM.DD.YYYY.xlsx/.xlsm".
- Prima del push fa SEMPRE git pull (anti-conflitto OneDrive / piu' PC).
- Nessuna finestra DOS: tutto l'output e' nel riquadro log.

Avvio: doppio clic su AVVIA_GUI.bat  (oppure  start "" pythonw app_dashboard_gui.py)
"""
import os
import re
import sys
import json
import time
import glob
import shutil
import subprocess
import threading
import urllib.request
import webbrowser
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl

# ----------------------------------------------------------------------
HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, "data")

PATTERN_OVERALL = "Sakarya Inspection Overall Status as of *.xlsx"
FOGLI_OVERALL = ["Dashboard", "Repair", "Rejection"]

PATTERN_3LPP = "3LPP Sakarya Inspection Overall Status as of *.xlsm"
FOGLI_3LPP = ["Production", "Lab Tests PPT", "Lab Tests Prod", "On hold",
              "Stripping- heating & Steel Dam.", "Raw Material"]

DASHBOARD_URL = "https://fanatics-hue.github.io/sakarya-dashboard/"
PAGE_3LPP_URL = "https://fanatics-hue.github.io/sakarya-dashboard/3lpp.html"
GH_API_REPO = "fanatics-hue/sakarya-dashboard"
CRUSCOTTO = os.path.join(os.path.dirname(HERE), "Cruscotto Workspace.hta")

C_IDLE = "#5a6472"
C_OK = "#1e8e3e"
C_WARN = "#e8a200"
C_ERR = "#c5221f"
C_BUSY = "#1a73e8"

_NO_WIN = 0x08000000 if os.name == "nt" else 0


def run(cmd):
    env = dict(os.environ, PYTHONUTF8="1", PYTHONIOENCODING="utf-8")
    p = subprocess.run(cmd, cwd=HERE, capture_output=True, text=True,
                       encoding="utf-8", errors="replace",
                       creationflags=_NO_WIN, env=env)
    out = (p.stdout or "") + (p.stderr or "")
    return p.returncode, out.strip()


def _api_get(url):
    req = urllib.request.Request(
        url, headers={"User-Agent": "sakarya-dashboard-gui",
                       "Accept": "application/vnd.github+json"})
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.loads(r.read().decode("utf-8"))


def verifica_deploy_pages(sha, log_fn, max_tries=12, wait_s=5):
    for _ in range(max_tries):
        time.sleep(wait_s)
        try:
            deployments = _api_get(
                "https://api.github.com/repos/%s/deployments?per_page=5" % GH_API_REPO)
            dep = next((d for d in deployments if d.get("sha") == sha), None)
            if not dep:
                continue
            statuses = _api_get(
                "https://api.github.com/repos/%s/deployments/%s/statuses"
                % (GH_API_REPO, dep["id"]))
            if not statuses:
                continue
            latest = statuses[0]
            if latest.get("state") in ("success", "failure", "error"):
                return latest["state"], latest.get("log_url")
        except Exception as e:
            log_fn("[avviso] verifica deploy: %s" % e)
    return "timeout", None


class SourcePanel:
    """Una sezione file+data per una delle due sorgenti (Overall Status / 3LPP)."""

    def __init__(self, parent, title, pattern, ext_label):
        self.pattern = pattern
        self.ext_label = ext_label
        self.excel_path = None
        self.mm = self.dd = self.yyyy = None
        self.done = False  # True se generato con successo in questa sessione

        self.frame = ttk.LabelFrame(parent, text=title, padding=8)

        rf = ttk.Frame(self.frame)
        rf.pack(fill="x", pady=(0, 4))
        ttk.Label(rf, text="File:", width=6).pack(side="left")
        self.lbl_file = ttk.Label(rf, text="(nessun file scelto)", foreground="#888")
        self.lbl_file.pack(side="left", fill="x", expand=True)
        ttk.Button(rf, text="Scegli file...", command=self.scegli_file).pack(side="right")

        rd = ttk.Frame(self.frame)
        rd.pack(fill="x")
        ttk.Label(rd, text="Data:", width=6).pack(side="left")
        self.ent_data = ttk.Entry(rd, width=14)
        self.ent_data.pack(side="left")
        ttk.Label(rd, text="(MM.DD.YYYY, dal nome file - modificabile)",
                  foreground="#888").pack(side="left", padx=8)

    def scegli_file(self):
        f = filedialog.askopenfilename(
            title="Scegli il file Excel",
            initialdir=HERE,
            filetypes=[(self.ext_label, "*." + self.ext_label.split(".")[-1].lower()), ("Tutti i file", "*.*")])
        if not f:
            return
        self.excel_path = f
        self.done = False
        self.lbl_file.configure(text=os.path.basename(f), foreground="#000")
        m = re.search(r"as of (\d{1,2})\.(\d{1,2})\.(\d{4})", os.path.basename(f))
        if m:
            self.mm, self.dd, self.yyyy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        else:
            now = datetime.now()
            self.mm, self.dd, self.yyyy = now.month, now.day, now.year
        self.ent_data.delete(0, "end")
        self.ent_data.insert(0, "%02d.%02d.%d" % (self.mm, self.dd, self.yyyy))

    def leggi_data_campo(self):
        s = self.ent_data.get().strip()
        m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
        if not m:
            return None
        return int(m.group(1)), int(m.group(2)), int(m.group(3))


class App:
    def __init__(self, root):
        self.root = root
        self.busy = False

        root.title("Sakarya Dashboard - Aggiornamento Settimanale")
        root.geometry("700x680")
        root.minsize(640, 600)

        self.banner = tk.Label(root, text="Pronto. Scegli uno o entrambi i file Excel per iniziare.",
                               bg=C_IDLE, fg="white", font=("Segoe UI", 12, "bold"),
                               anchor="w", padx=14, pady=12)
        self.banner.pack(fill="x")

        body = ttk.Frame(root, padding=12)
        body.pack(fill="both", expand=True)

        self.p_overall = SourcePanel(body, "Overall Status  (file .xlsx)", PATTERN_OVERALL, ".xlsx")
        self.p_overall.frame.pack(fill="x", pady=(0, 8))
        self.p_3lpp = SourcePanel(body, "3LPP Inspection  (file .xlsm)", PATTERN_3LPP, ".xlsm")
        self.p_3lpp.frame.pack(fill="x", pady=(0, 10))

        rb = ttk.Frame(body)
        rb.pack(fill="x", pady=(0, 10))
        self.btn_gen = ttk.Button(rb, text="1.  Genera e verifica", command=self.t_genera)
        self.btn_gen.pack(side="left")
        self.btn_push = ttk.Button(rb, text="2.  Pubblica su GitHub",
                                   command=self.t_pubblica, state="disabled")
        self.btn_push.pack(side="left", padx=8)

        ttk.Label(body, text="Dettagli:").pack(anchor="w")
        wrap = ttk.Frame(body)
        wrap.pack(fill="both", expand=True)
        self.txt = tk.Text(wrap, height=16, wrap="word", font=("Consolas", 9),
                           bg="#1e1e1e", fg="#d4d4d4", insertbackground="white")
        sb = ttk.Scrollbar(wrap, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set, state="disabled")
        self.txt.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        rbot = ttk.Frame(body)
        rbot.pack(fill="x", pady=(10, 0))
        ttk.Button(rbot, text="Apri Overall Status",
                   command=lambda: webbrowser.open(DASHBOARD_URL)).pack(side="left")
        ttk.Button(rbot, text="Apri 3LPP Inspection",
                   command=lambda: webbrowser.open(PAGE_3LPP_URL)).pack(side="left", padx=8)
        ttk.Button(rbot, text="Apri cartella",
                   command=lambda: os.startfile(HERE)).pack(side="left")
        if os.path.exists(CRUSCOTTO):
            ttk.Button(rbot, text="Cruscotto IQS",
                       command=lambda: os.startfile(CRUSCOTTO)).pack(side="right")

        self.log("Pronto. Repository: fanatics-hue/sakarya-dashboard (branch main).")
        self.log("Scegli il file 'Overall Status' e/o il file '3LPP' (anche uno solo), poi '1. Genera e verifica'.")

    # ------------------------------------------------------------------ utils
    def log(self, msg):
        self.txt.configure(state="normal")
        self.txt.insert("end", msg + "\n")
        self.txt.see("end")
        self.txt.configure(state="disabled")

    def set_banner(self, text, color):
        self.banner.configure(text=text, bg=color)

    def lock(self, on):
        self.busy = on
        st = "disabled" if on else "normal"
        self.btn_gen.configure(state=st)

    # --------------------------------------------------------------- genera
    def t_genera(self):
        threading.Thread(target=self.genera, daemon=True).start()

    def genera(self):
        try:
            self.lock(True)
            self.btn_push.configure(state="disabled")
            self.set_banner("Verifica e generazione in corso...", C_BUSY)

            if not self.p_overall.excel_path and not self.p_3lpp.excel_path:
                self.set_banner("STOP: scegli almeno un file Excel.", C_ERR)
                self.log("[ERRORE] Nessun file scelto.")
                return

            ok_overall = self._genera_overall() if self.p_overall.excel_path else None
            ok_3lpp = self._genera_3lpp() if self.p_3lpp.excel_path else None

            risultati = [r for r in (ok_overall, ok_3lpp) if r is not None]
            if not risultati:
                return
            if all(risultati):
                fatti = []
                if ok_overall:
                    fatti.append("Overall Status")
                if ok_3lpp:
                    fatti.append("3LPP")
                self.set_banner("OK - pronto (%s). Controlla i KPI sopra, poi '2. Pubblica'." % " + ".join(fatti), C_OK)
                self.btn_push.configure(state="normal")
            elif any(risultati):
                self.set_banner("Uno dei due e' fallito - controlla il log. Puoi comunque pubblicare l'altro.", C_WARN)
                self.btn_push.configure(state="normal")
            else:
                self.set_banner("STOP: entrambe le generazioni sono fallite.", C_ERR)
        finally:
            self.lock(False)

    def _genera_overall(self):
        p = self.p_overall
        self.log("=" * 56)
        self.log("OVERALL STATUS")
        d = p.leggi_data_campo()
        if not d:
            self.log("[ERRORE] Data non valida: '%s'" % p.ent_data.get())
            return False
        p.mm, p.dd, p.yyyy = d

        try:
            wb = openpyxl.load_workbook(p.excel_path, read_only=True)
            mancanti = [s for s in FOGLI_OVERALL if s not in wb.sheetnames]
            wb.close()
        except Exception as e:
            self.log("[ERRORE] Impossibile aprire l'Excel: %s" % e)
            return False
        if mancanti:
            self.log("[ERRORE] Mancano i fogli: %s" % ", ".join(mancanti))
            return False
        self.log("Fogli richiesti presenti: %s" % ", ".join(FOGLI_OVERALL))

        target_name = "Sakarya Inspection Overall Status as of %02d.%02d.%d.xlsx" % (p.mm, p.dd, p.yyyy)
        target_path = os.path.join(HERE, target_name)
        if os.path.abspath(p.excel_path) != os.path.abspath(target_path):
            shutil.copy2(p.excel_path, target_path)
            self.log("File copiato come: %s" % target_name)
        else:
            os.utime(target_path, None)
        p.excel_path = target_path

        self.log("Genero i dati (genera_dati.py)...")
        rc, out = run([sys.executable, os.path.join(HERE, "genera_dati.py")])
        for line in out.splitlines():
            self.log("   " + line)
        if rc != 0:
            self.log("[ERRORE] Generazione Overall Status fallita.")
            return False
        p.done = True
        return True

    def _genera_3lpp(self):
        p = self.p_3lpp
        self.log("=" * 56)
        self.log("3LPP INSPECTION")
        d = p.leggi_data_campo()
        if not d:
            self.log("[ERRORE] Data non valida: '%s'" % p.ent_data.get())
            return False
        p.mm, p.dd, p.yyyy = d

        try:
            wb = openpyxl.load_workbook(p.excel_path, read_only=True, keep_vba=True)
            mancanti = [s for s in FOGLI_3LPP if s not in wb.sheetnames]
            wb.close()
        except Exception as e:
            self.log("[ERRORE] Impossibile aprire l'Excel: %s" % e)
            return False
        if mancanti:
            self.log("[ERRORE] Mancano i fogli: %s" % ", ".join(mancanti))
            return False
        self.log("Fogli richiesti presenti: %s" % ", ".join(FOGLI_3LPP))

        target_name = "3LPP Sakarya Inspection Overall Status as of %02d.%02d.%d.xlsm" % (p.mm, p.dd, p.yyyy)
        target_path = os.path.join(HERE, target_name)
        if os.path.abspath(p.excel_path) != os.path.abspath(target_path):
            shutil.copy2(p.excel_path, target_path)
            self.log("File copiato come: %s" % target_name)
        else:
            os.utime(target_path, None)
        p.excel_path = target_path

        self.log("Genero la pagina (genera_dati_3lpp.py)...")
        rc, out = run([sys.executable, os.path.join(HERE, "genera_dati_3lpp.py")])
        for line in out.splitlines():
            self.log("   " + line)
        if rc != 0:
            self.log("[ERRORE] Generazione 3LPP fallita.")
            return False
        p.done = True
        return True

    # ----------------------------------------------------------------- pubblica
    def t_pubblica(self):
        fatti = []
        if self.p_overall.done:
            fatti.append("Overall Status (%02d.%02d.%d)" % (self.p_overall.mm, self.p_overall.dd, self.p_overall.yyyy))
        if self.p_3lpp.done:
            fatti.append("3LPP (%02d.%02d.%d)" % (self.p_3lpp.mm, self.p_3lpp.dd, self.p_3lpp.yyyy))
        if not messagebox.askyesno(
                "Conferma pubblicazione",
                "Pubblicare su GitHub?\n\n%s\n\nIl sito online verra' aggiornato per tutti i colleghi." % "\n".join(fatti)):
            return
        threading.Thread(target=self.pubblica, daemon=True).start()

    def pubblica(self):
        try:
            self.lock(True)
            self.btn_push.configure(state="disabled")
            self.set_banner("Pubblicazione in corso...", C_BUSY)

            rc, _ = run(["git", "--version"])
            if rc != 0:
                self.set_banner("STOP: Git non installato.", C_ERR)
                self.log("[ERRORE] Git non trovato. Installa da https://git-scm.com")
                return

            rc, branch = run(["git", "rev-parse", "--abbrev-ref", "HEAD"])
            branch = branch.strip() or "main"

            self.log("-" * 56)
            self.log("git pull origin %s ..." % branch)
            rc, out = run(["git", "pull", "--no-edit", "origin", branch])
            self.log(out or "(ok)")
            if rc != 0:
                self.set_banner("STOP: pull fallito (conflitto o rete).", C_ERR)
                self.log("[ERRORE] Risolvi il conflitto/connessione e riprova.")
                return

            run(["git", "add", "-A"])
            rc, _ = run(["git", "diff", "--cached", "--quiet"])
            has_changes = (rc != 0)

            if has_changes:
                parti = []
                if self.p_overall.done:
                    parti.append("Overall Status %02d.%02d.%d" % (self.p_overall.mm, self.p_overall.dd, self.p_overall.yyyy))
                if self.p_3lpp.done:
                    parti.append("3LPP %02d.%02d.%d" % (self.p_3lpp.mm, self.p_3lpp.dd, self.p_3lpp.yyyy))
                msg = "Aggiornamento: " + " + ".join(parti) if parti else "Aggiornamento dashboard"
                rc, out = run(["git", "commit", "-m", msg])
                self.log(out)
            else:
                # Nessuna modifica nuova rispetto all'ultimo commit locale.
                # ATTENZIONE: questo NON significa che GitHub sia aggiornato -
                # un push precedente potrebbe essere fallito o essere stato
                # interrotto (rete, credenziali, OneDrive che sincronizza la
                # cartella .git), lasciando commit locali mai arrivati online.
                # Controlliamo quindi se HEAD e' avanti rispetto a origin/<branch>
                # (il pull appena fatto ha gia' aggiornato il ref origin/<branch>).
                # Dettaglio: LL-024 nel registro Lessons Learned Order 45650.
                rc, ahead_out = run(["git", "rev-list", "--count",
                                     "origin/%s..HEAD" % branch])
                try:
                    ahead_n = int(ahead_out.strip())
                except (ValueError, TypeError):
                    ahead_n = 0

                if ahead_n == 0:
                    self.set_banner("Niente da pubblicare: gia' aggiornata.", C_WARN)
                    self.log("Nessuna modifica rispetto a GitHub.")
                    return
                else:
                    self.log("Nessuna modifica nuova da committare, ma ho trovato "
                              "%d commit locali non ancora su GitHub "
                              "(probabile push precedente non riuscito) "
                              "- li pubblico ora." % ahead_n)

            self.log("git push origin %s ..." % branch)
            rc, out = run(["git", "push", "origin", branch])
            self.log(out or "(ok)")
            if rc != 0:
                self.set_banner("STOP: push fallito (credenziali/rete).", C_ERR)
                self.log("[ERRORE] Controlla login GitHub e connessione.")
                return

            rc, sha = run(["git", "rev-parse", "HEAD"])
            sha = sha.strip()
            self.set_banner("Push ok. Verifico la pubblicazione online...", C_BUSY)
            self.log("-" * 56)
            self.log("Verifico che GitHub Pages abbia pubblicato (fino a 1 min)...")
            stato, log_url = verifica_deploy_pages(sha, self.log)

            if stato == "success":
                self.set_banner("FATTO! Pubblicato (CTRL+F5 per vedere).", C_OK)
                self.log("COMPLETATO.")
                if self.p_overall.done:
                    self.log("  Overall Status: %s" % DASHBOARD_URL)
                if self.p_3lpp.done:
                    self.log("  3LPP: %s" % PAGE_3LPP_URL)
            elif stato == "failure":
                self.set_banner("ATTENZIONE: push ok ma la pubblicazione e' FALLITA.", C_ERR)
                self.log("[ERRORE] Il push e' arrivato su GitHub ma la build di Pages e' fallita.")
                self.log("Come risolvere:")
                self.log("  1) Apri: %s e clicca 'Re-run failed jobs'." % log_url)
                self.log("  2) Se non basta, premi di nuovo '2. Pubblica su GitHub'.")
            else:
                self.set_banner("Push ok, pubblicazione ancora in corso.", C_WARN)
                self.log("Non sono riuscito a confermare la pubblicazione in tempo.")
                self.log("Aspetta 1-2 minuti e ricarica con CTRL+F5.")
        finally:
            self.lock(False)


def main():
    root = tk.Tk()
    try:
        ttk.Style().theme_use("vista")
    except Exception:
        pass
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
